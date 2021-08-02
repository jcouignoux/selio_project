from django.views.generic import View
from django.http import HttpResponse
from django.forms.models import model_to_dict
import xlrd
import openpyxl
import xlsxwriter, xlrd
from io import BytesIO, StringIO

from wsgiref.util import FileWrapper


from .models import Author, Publisher, Categorie, Ouvrage

class xlsx():

    def importXLSX(self, file):
        wb = xlrd.open_workbook(file_contents=file.read())
        #wb = openpyxl.load_workbook(file, data_only=True)
        # # Table Catégorie
        # catSheet = wb.sheet_by_name("categorie")
        # for r in range(2, catSheet.max_row):
        #     cat = catSheet.cell(rowx=r, colx=2).value
        #     categorie = Categorie()
        #     categorie.name = cat
        #     categorie.save()
# 
        # pubSheet = wb.sheet_by_name("editeur")
        # for r in range(2, pubSheet.max_row):
        #     pubName = pubSheet.cell(rowx=r, colx=2).value
        #     pubMarge = pubSheet.cell(rowx=r, colx=3).value
        #     publisher = Publisher()
        #     publisher.name = pubName
        #     publisher.marge = pubMarge
        #     publisher.save()
# 
        # autSheet = wb.sheet_by_name("auteur")
        # for r in range(2, autSheet.max_row):
        #     autName = autSheet.cell(rowx=r, colx=2).value
        #     autForname = autSheet.cell(rowx=r, colx=3).value
        #     author = Author()
        #     author.name = autName
        #     author.forname = autForname
        #     author.save()

        ouvSheet = wb.get_sheet_by_name("ouvrage")
        for r in range(2, ouvSheet.max_row):
            # with transaction.atomic():
            ouvrage = Ouvrage()
            ouvRef = ouvSheet.cell(rowx=r, colx=2).value
            ouvrage.reference=ouvRef
            ouvrage.save()
            ouvTitle = ouvSheet.cell(rowx=r, colx=3).value
            ouvrage.title=ouvTitle
            ouvPub = ouvSheet.cell(rowx=r, colx=4).value
            print(ouvTitle)
            pubObj = Publisher.objects.filter(name=ouvPub).first()
            if pubObj != None:
                ouvrage.editeurs.add(pubObj.id)
            ouvAuthors = ouvSheet.cell(rowx=r, colx=5).value
            for author in ouvAuthors.split(' et '):
                autName = author.split(', ')[0]
                autForname = author.split(', ')[1]
                if autForname != "":
                    autObj = Author.objects.filter(name=autName, forname=autForname).first()
                else:
                    autObj = Author.objects.filter(name=autName).first()
                ouvrage.auteurs.add(autObj.id)
            ouvCats = ouvSheet.cell(rowx=r, colx=6).value
            catObj = Categorie.objects.filter(name=ouvCats).first()
            if catObj != None:
                ouvrage.categories.add(catObj.id)
            ouvPar = ouvSheet.cell(rowx=r, colx=7).value
            ouvrage.publication=ouvPar
            ouvPrice = ouvSheet.cell(rowx=r, colx=8).value
            ouvrage.price=ouvPrice
            ouvStock = ouvSheet.cell(rowx=r, colx=11).value
            if ouvStock == "":
                ouvrage.stock=0
            else:
                ouvrage.stock=ouvStock
            ouvPic = ouvSheet.cell(rowx=r, colx=12).value
            if ouvPic == "":
                ouvrage.picture = None
            else:
                ouvrage.picture=ouvPic
            ouvNote = ouvSheet.cell(rowx=r, colx=13).value
            ouvrage.note=ouvNote
            ouvrage.save()
        return file

def WriteToExcel(dict, title, name):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    sheet1 = workbook.add_worksheet(name)
    sheet1.autofilter(0, 0, 0, 10)
    sheet1.set_column('A:A', 18)
    sheet1.set_column('B:B', 18)
    sheet1.set_column('C:C', 63)
    sheet1.set_column('D:D', 28)
    sheet1.set_column('E:E', 40)
    sheet1.set_column('F:F', 8)
    sheet1.set_column('G:G', 10)
    sheet1.set_column('H:H', 10)
    sheet1.set_column('I:I', 20)
    sheet1.set_column('J:J', 8)
    sheet1.set_column('K:K', 30)

    pairFmt = workbook.add_format({'bg_color': '#E0E0E0'})
    impairFmt = workbook.add_format({'bg_color': '#C0C0C0'})
    titreFmt = workbook.add_format({'font_color': 'white', 'bg_color': '#003366', 'bold' : True})
    pprixFmt = workbook.add_format({'num_format': '# ##0.00 €', 'bg_color': '#E0E0E0'})
    iprixFmt = workbook.add_format({'num_format': '# ##0.00 €', 'bg_color': '#C0C0C0'})
    prefFmt = workbook.add_format({'num_format': '0', 'bg_color': '#E0E0E0', 'align': 'left'})
    irefFmt = workbook.add_format({'num_format': '0', 'bg_color': '#C0C0C0', 'align': 'left'})
    pdateFmt = workbook.add_format({'num_format': 'yyyy/mm/dd', 'bg_color': '#E0E0E0', 'align': 'left'})
    idateFmt = workbook.add_format({'num_format': 'yyyy/mm/dd', 'bg_color': '#C0C0C0', 'align': 'left'})
    stockFmt = workbook.add_format({'font_color': 'red'})

    row=0
    col=0
    for line in dict.iterator():
        for cel in ('date', 'reference', 'title', 'auteurs', 'editeurs', 'price', 'catPrice', 'payment', 'fournisseur', 'quantity', 'comment'):
            if row == 0:
                sheet1.write(row, col, cel, titreFmt)
                col += 1
            else:
                sheet1.conditional_format('H1:H500', {'type': 'cell', 'criteria': '<', 'value': 5, 'format': stockFmt})
                if (row % 2) == 0:
                    Fmt = workbook.add_format({'bg_color': '#E0E0E0'})
                else:
                    Fmt = workbook.add_format({'bg_color': '#C0C0C0'})
                if cel == 'auteurs':
                    auteurs = []
                    for auteur in line.__dict__[cel]:
                        # auteurs.append(str(auteurs)+', '+str(auteurs))
                        pass
                    sheet1.write(row, col, str(' et '.join(auteurs)), Fmt)
                elif cel == 'editeurs':
                    try:
                        sheet1.write(row, col, str(line.__dict__[cel]), Fmt)
                    except:
                        pass
                if cel == 'price':
                    if (row % 2) == 0:
                        sheet1.write(row, col, line.__dict__[cel], pprixFmt)
                    else:
                        sheet1.write(row, col, line.__dict__[cel], iprixFmt)
                elif cel == 'reference':
                    if (row % 2) == 0:
                        sheet1.write(row, col, line.__dict__[cel], prefFmt)
                    else:
                        sheet1.write(row, col, line.__dict__[cel], irefFmt)
                elif cel == 'date':
                    if (row % 2) == 0:
                        sheet1.write(row, col, line.__dict__[cel], pdateFmt)
                    else:
                        sheet1.write(row, col, line.__dict__[cel], idateFmt)
                else:
                    sheet1.write(row, col, line.__dict__[cel], Fmt)
                col += 1
        row +=1
        col=0

    workbook.close()
    xlsx_data = output.getvalue()
    # xlsx_data contains the Excel file
    return xlsx_data

def exportXLSX(dict, title, name):
    downFile = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    downFile['Content-Disposition'] = 'attachment; filename="Comptes.xlsx"'
    xlsx_data = WriteToExcel(dict, title, name)
    downFile.write(xlsx_data)
    return downFile 